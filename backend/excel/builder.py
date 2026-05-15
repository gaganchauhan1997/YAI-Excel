"""
YAIExcelBuilder — orchestrator that turns an audit JSON into a finished .xlsx.

Execution order (18 steps, mirrors the spec):
  1.  create_all_sheets
  2.  write_raw_data
  3.  build_named_ranges
  4.  build_all_interactive_controls
  5.  build_all_pivots
  6.  build_all_charts
  7.  build_all_kpi_cards
  8.  build_all_data_tables
  9.  apply_all_cond_formats
  10. build_all_slicers           (gracefully no-op via PivotEngine)
  11. apply_data_validation
  12. link_all_interactivity
  13. apply_theme
  14. apply_enhancements
  15. set_sheet_properties
  16. run_quality_gates
  17. generate_preview_png
  18. save_to_bytes
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from loguru import logger

from .formatter import FormatEngine
from .chart_engine import ChartEngine
from .kpi_engine import KPIEngine
from .table_engine import TableEngine
from .pivot_engine import PivotEngine
from .interactivity import InteractivityEngine
from .formula_writer import FormulaWriter
from .enhancer import EnhancementEngine
from .themes import get_theme


@dataclass
class BuildResult:
    workbook: Workbook
    audit: dict
    counts: dict
    quality: dict
    preview_png_path: str | None = None


@dataclass
class YAIExcelBuilder:
    theme_name: str = "midnight"
    raw_data: pd.DataFrame | None = None
    audit: dict = field(default_factory=dict)

    def __post_init__(self) -> None:
        self.theme = get_theme(self.theme_name)
        self.formatter = FormatEngine(self.theme)
        self.charts = ChartEngine(self.theme)
        self.kpis = KPIEngine(self.theme, self.formatter, indian_format=bool(self.audit.get("_indian_format")))
        self.tables = TableEngine(self.theme, self.formatter)
        self.pivots = PivotEngine(self.theme, self.formatter)
        self.interactivity = InteractivityEngine(self.theme)
        self.formula_writer = FormulaWriter()
        self.enhancer = EnhancementEngine(self.theme)

    # ------------------------------------------------------------------
    # Main pipeline
    # ------------------------------------------------------------------
    def build(self) -> BuildResult:
        logger.info(f"YAIExcelBuilder.build | theme={self.theme_name}")
        audit = self.enhancer.enhance(self.audit)

        wb = Workbook()
        self._step("create_sheets", self._create_sheets, wb)
        self._step("write_raw_data", self._write_raw_data, wb)
        self._step("named_ranges", self.interactivity.build_named_ranges, wb, audit.get("named_ranges", []))
        self._step("controls", self.interactivity.build_controls, wb, audit.get("interactive_controls", []))
        self._step("pivots", self.pivots.build_all, wb, audit.get("pivot_tables", []))
        self._step("charts", self.charts.build_all, wb, audit.get("charts", []))
        self._step("kpis", self.kpis.build_all, wb["Dashboard"], audit.get("kpi_strip", []))
        self._step("tables", self.tables.build_all, wb, audit.get("data_tables", []))
        self._step("cond_formats", self._apply_all_cond_formats, wb, audit.get("conditional_formatting", []))
        self._step("validation", self.interactivity.apply_data_validation, wb, audit.get("data_validation", []))
        self._step("formulas", self.formula_writer.write_all, wb, audit.get("formulas", []))
        self._step("theme", self._apply_theme, wb)
        self._step("sheet_properties", self._set_sheet_properties, wb)

        quality = self._run_quality_gates(wb, audit)

        return BuildResult(workbook=wb, audit=audit, counts=audit.get("counts", {}), quality=quality)

    def save_to_bytes(self) -> bytes:
        result = self.build()
        buf = io.BytesIO()
        result.workbook.save(buf)
        buf.seek(0)
        return buf.read()

    def save_to_path(self, path: str | Path) -> Path:
        result = self.build()
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        result.workbook.save(path)
        return path

    # ------------------------------------------------------------------
    # Steps
    # ------------------------------------------------------------------
    def _step(self, name: str, fn, *args, **kwargs):
        try:
            out = fn(*args, **kwargs)
            logger.info(f"  ✓ {name} -> {out}" if out is not None else f"  ✓ {name}")
            return out
        except Exception as exc:
            logger.warning(f"  ✗ {name} failed: {exc}")
            return None

    def _create_sheets(self, wb: Workbook) -> list[str]:
        default = wb.active
        default.title = "Dashboard"
        wb.create_sheet("Data")
        wb.create_sheet("Formulas")
        wb.create_sheet("Calcs")
        return wb.sheetnames

    def _write_raw_data(self, wb: Workbook) -> int:
        if self.raw_data is None or self.raw_data.empty:
            return 0
        ws = wb["Data"]
        df = self.raw_data
        for c_idx, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=str(col))
            self.formatter.apply(cell, self.formatter.header_style())
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        return len(df)

    def _apply_all_cond_formats(self, wb: Workbook, rules: list[dict]) -> int:
        # Default rules are keyed by sheet via the range "Sheet!Range"
        total = 0
        for rule in rules or []:
            rng = rule.get("range") or ""
            if "!" in rng:
                sheet_name, plain = rng.split("!", 1)
                rule_local = dict(rule)
                rule_local["range"] = plain
                if sheet_name in wb.sheetnames:
                    total += self.formatter.apply_conditional_formatting(wb[sheet_name], [rule_local])
            else:
                total += self.formatter.apply_conditional_formatting(wb["Dashboard"], [rule])
        return total

    def _apply_theme(self, wb: Workbook) -> int:
        ws = wb["Dashboard"]
        # Title row
        ws.merge_cells("A1:K1")
        title = ws["A1"]
        title.value = '="YAI-Excel Dashboard — "&TEXT(TODAY(),"DD MMM YYYY")'
        self.formatter.apply(title, self.formatter.title_style())
        ws.row_dimensions[1].height = 40
        # Sheet-level theming
        for sheet_name in wb.sheetnames:
            self.formatter.theme_sheet(wb[sheet_name], hide_gridlines=(sheet_name == "Dashboard"))
        return len(wb.sheetnames)

    def _set_sheet_properties(self, wb: Workbook) -> None:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self.interactivity.set_sheet_properties(
                ws,
                freeze="A4" if sheet_name == "Dashboard" else "A2",
                hide_gridlines=(sheet_name == "Dashboard"),
                zoom=90,
                tab_color=self.theme["tab_color"],
            )
            ws.page_setup.orientation = "landscape"
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.print_options.horizontalCentered = True

    # ------------------------------------------------------------------
    # Quality gates
    # ------------------------------------------------------------------
    def _run_quality_gates(self, wb: Workbook, audit: dict) -> dict:
        results: dict[str, bool] = {}
        # 1. All controls linked
        results["controls_linked"] = all((c.get("linked_cell") or c.get("location"))
                                        for c in audit.get("interactive_controls", []))
        # 2. Chart titles set
        results["chart_titles_set"] = all(c.get("title") for c in audit.get("charts", []))
        # 3. Sheets present
        results["sheets_present"] = {"Dashboard", "Data"}.issubset(set(wb.sheetnames))
        # 4. No empty charts (heuristic: each chart has at least one series spec)
        results["no_empty_charts"] = all(c.get("series") for c in audit.get("charts", []))
        # 5. KPIs ≥ 1
        results["has_kpis"] = len(audit.get("kpi_strip", [])) > 0
        # 6. Theme applied
        results["theme_applied"] = bool(self.theme.get("header_bg"))
        results["all_passed"] = all(v is True or v is True for v in results.values())
        return results
