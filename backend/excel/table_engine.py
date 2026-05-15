"""
TableEngine — paints styled data tables on the Dashboard or Data sheet.

Wraps the data in an Excel Table (formal `ListObject`) so users get
filter/sort headers + built-in striping + structured references.
"""
from __future__ import annotations

from dataclasses import dataclass

from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.table import Table, TableStyleInfo
from loguru import logger

from .formatter import FormatEngine


@dataclass
class TableEngine:
    theme: dict
    formatter: FormatEngine

    def build_all(self, wb, tables: list[dict]) -> int:
        built = 0
        for idx, t in enumerate(tables or []):
            try:
                self.build_one(wb, idx, t)
                built += 1
            except Exception as exc:
                logger.warning(f"table {t.get('id','?')} failed: {exc}")
        return built

    def build_one(self, wb, idx: int, t: dict) -> None:
        sheet_name = t.get("sheet") or ("Data" if "Data" in wb.sheetnames else "Dashboard")
        ws = wb[sheet_name]

        anchor = t.get("location") or "A1"
        col_letter, row = coordinate_from_string(anchor)
        col = column_index_from_string(col_letter)

        headers = t.get("headers") or []
        row_count = max(t.get("row_count") or 1, 1)

        # Write headers if not present
        for c_off, header in enumerate(headers):
            cell = ws.cell(row=row, column=col + c_off, value=header)
            self.formatter.apply(cell, self.formatter.header_style())

        end_col = col + max(len(headers), 1) - 1
        end_row = row + row_count
        ref = f"{get_column_letter(col)}{row}:{get_column_letter(end_col)}{end_row}"

        try:
            tbl = Table(displayName=f"Tbl_{t.get('id', f'table_{idx+1}')}", ref=ref)
            tbl.tableStyleInfo = TableStyleInfo(
                name=t.get("table_style") or "TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(tbl)
        except Exception as exc:
            logger.warning(f"could not register Excel Table for {t.get('id','?')}: {exc}")
            # fall back to manual striping
            self.formatter.stripe_table(ws, row + 1, end_row, col, end_col)

        if t.get("has_totals_row"):
            total_row = end_row + 1
            ws.cell(row=total_row, column=col, value="Total").font = self.formatter.header_style()["font"]
            for c_off in range(1, len(headers)):
                col_letter = get_column_letter(col + c_off)
                ws.cell(row=total_row, column=col + c_off,
                        value=f"=SUM({col_letter}{row+1}:{col_letter}{end_row})")
