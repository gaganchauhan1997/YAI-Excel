"""
InteractivityEngine — wires up dropdowns, button groups, tab cells, named
ranges, and data validation so the dashboard responds to user input.
"""
from __future__ import annotations

from dataclasses import dataclass

from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from loguru import logger


@dataclass
class InteractivityEngine:
    theme: dict

    # ------------------------------------------------------------------
    # Named ranges
    # ------------------------------------------------------------------
    def build_named_ranges(self, wb, ranges: list[dict]) -> int:
        built = 0
        for r in ranges or []:
            name = r.get("name")
            refers_to = r.get("refers_to")
            if not name or not refers_to:
                continue
            try:
                wb.defined_names[name] = DefinedName(name=name, attr_text=refers_to)
                built += 1
            except Exception as exc:
                logger.warning(f"named range {name} failed: {exc}")
        return built

    # ------------------------------------------------------------------
    # Data validation (dropdowns, numeric, date)
    # ------------------------------------------------------------------
    def apply_data_validation(self, wb, rules: list[dict]) -> int:
        applied = 0
        for rule in rules or []:
            rng = rule.get("range")
            if not rng:
                continue
            try:
                ws = wb["Dashboard"] if "Dashboard" in wb.sheetnames else wb.active
                rtype = (rule.get("type") or "list").lower()
                if rtype == "list":
                    formula = rule.get("formula") or '"Q1,Q2,Q3,Q4"'
                    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
                elif rtype == "whole":
                    dv = DataValidation(type="whole", operator="greaterThan", formula1="0")
                elif rtype == "decimal":
                    dv = DataValidation(type="decimal", operator="greaterThan", formula1="0")
                elif rtype == "date":
                    dv = DataValidation(type="date", operator="greaterThan", formula1="DATE(2000,1,1)")
                else:
                    continue
                dv.error = rule.get("error_message") or "Invalid input"
                dv.errorTitle = "YAI-Excel"
                dv.add(rng)
                ws.add_data_validation(dv)
                applied += 1
            except Exception as exc:
                logger.warning(f"data validation failed: {exc}")
        return applied

    # ------------------------------------------------------------------
    # Interactive controls — main dropdown + filter buttons
    # ------------------------------------------------------------------
    def build_controls(self, wb, controls: list[dict]) -> int:
        built = 0
        ws = wb["Dashboard"] if "Dashboard" in wb.sheetnames else wb.active
        for c in controls or []:
            ctype = (c.get("type") or "dropdown").lower()
            location = c.get("location") or "B1"
            options = c.get("options") or []
            if ctype == "dropdown" and options:
                formula = '"' + ",".join(str(o).replace(",", "") for o in options) + '"'
                dv = DataValidation(type="list", formula1=formula, allow_blank=True)
                dv.add(location)
                ws.add_data_validation(dv)
                ws[location] = options[0] if options else ""
                built += 1
            elif ctype == "checkbox":
                ws[location] = "FALSE"
                dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
                dv.add(location)
                ws.add_data_validation(dv)
                built += 1
            else:
                ws[location] = options[0] if options else ""
                built += 1
        return built

    # ------------------------------------------------------------------
    # Dashboard polish: freeze panes, hidden gridlines, print area
    # ------------------------------------------------------------------
    def set_sheet_properties(self, ws, *, freeze: str = "A4",
                             hide_gridlines: bool = True,
                             zoom: int = 90,
                             tab_color: str | None = None) -> None:
        ws.freeze_panes = freeze
        ws.sheet_view.showGridLines = not hide_gridlines
        ws.sheet_view.zoomScale = zoom
        if tab_color:
            ws.sheet_properties.tabColor = tab_color
