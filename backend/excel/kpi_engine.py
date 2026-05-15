"""
KPIEngine — paints KPI cards on the Dashboard sheet.

Each KPI card is a 4-row x 3-col block:
    row 0: label
    row 1: value (large font, formula-driven)
    row 2: trend arrow (formula-driven, conditional-colored)
"""
from __future__ import annotations

from dataclasses import dataclass

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from loguru import logger

from .formatter import FormatEngine


@dataclass
class KPIEngine:
    theme: dict
    formatter: FormatEngine

    def build_all(self, ws, kpis: list[dict]) -> int:
        built = 0
        for idx, kpi in enumerate(kpis or []):
            try:
                self._paint_card(ws, idx, kpi)
                built += 1
            except Exception as exc:
                logger.warning(f"kpi {kpi.get('id','?')} failed: {exc}")
        return built

    def _paint_card(self, ws, idx: int, kpi: dict) -> None:
        anchor = kpi.get("location") or self._auto_location(idx)
        col_letter, row = coordinate_from_string(anchor)
        col = column_index_from_string(col_letter)

        # Card spans 3 columns x 3 rows
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 2)
        ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 2)

        bg = (kpi.get("bg_color") or "#" + self.theme["kpi_bg"]).lstrip("#")
        fg = (kpi.get("text_color") or "#" + self.theme["kpi_fg"]).lstrip("#")

        label_cell = ws.cell(row=row, column=col, value=kpi.get("label") or kpi.get("id", "KPI"))
        label_cell.font = Font(name=self.theme["font_name"], color=fg, bold=True, size=11)
        label_cell.fill = PatternFill("solid", fgColor=bg)
        label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        value_cell = ws.cell(row=row + 1, column=col)
        formula = kpi.get("formula") or kpi.get("value") or ""
        if formula and not str(formula).startswith("="):
            formula = f"={formula}" if any(c.isalpha() or c == "(" for c in str(formula)) else formula
        value_cell.value = formula or kpi.get("value")
        value_cell.font = Font(name=self.theme["font_name"], color=fg, bold=True, size=22)
        value_cell.fill = PatternFill("solid", fgColor=bg)
        value_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        value_cell.number_format = '#,##0.00;[Red]-#,##0.00'

        trend_cell = ws.cell(row=row + 2, column=col)
        trend_cell.value = kpi.get("trend_formula") or self._default_trend(kpi)
        trend_cell.font = Font(name=self.theme["font_name"], color=self.theme["kpi_trend_up"], bold=True, size=11)
        trend_cell.fill = PatternFill("solid", fgColor=bg)
        trend_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Set row heights for visual breathing room
        ws.row_dimensions[row].height = 22
        ws.row_dimensions[row + 1].height = 36
        ws.row_dimensions[row + 2].height = 18

    @staticmethod
    def _auto_location(idx: int) -> str:
        # Strip across row 3, 4 cols wide per card, with 1-col gap
        base_col = 2 + idx * 4
        return f"{get_column_letter(base_col)}3"

    @staticmethod
    def _default_trend(kpi: dict) -> str:
        # Generic placeholder; the real formula would reference previous-period cell
        trend = (kpi.get("trend") or "neutral").lower()
        glyph = {"up": "▲", "down": "▼"}.get(trend, "→")
        return f'="{glyph} vs last period"'
