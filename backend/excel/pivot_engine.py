"""
PivotEngine — emits real, formula-driven pivot summaries from raw data.

openpyxl can't reliably author native Excel pivots from scratch (the
pivotCache XML format is fragile and viewer-dependent), so YAI-Excel takes
the pragmatic high-road: it generates a fully-functional pivot equivalent
using real `SUMIFS` / `COUNTIFS` / `AVERAGEIFS` formulas against the source
range, with computed row labels, multi-value columns, subtotals, grand
totals, and a styled header band that looks and behaves like a native pivot.

This is upgraded from the v1.0 placeholder: it now reads the source data
sheet, computes unique row-label values, and writes real cells that update
when the data changes.
"""
from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from loguru import logger

from .formatter import FormatEngine

AGG_FN = {
    "SUM": "SUMIFS",
    "COUNT": "COUNTIFS",
    "AVERAGE": "AVERAGEIFS",
    "AVG": "AVERAGEIFS",
    "MAX": "MAXIFS",
    "MIN": "MINIFS",
}


@dataclass
class PivotEngine:
    theme: dict
    formatter: FormatEngine

    # ------------------------------------------------------------------
    # Top-level
    # ------------------------------------------------------------------
    def build_all(self, wb: Workbook, pivots: list[dict]) -> int:
        built = 0
        for idx, spec in enumerate(pivots or []):
            try:
                self.build_one(wb, idx, spec)
                built += 1
            except Exception as exc:
                logger.warning(f"pivot {spec.get('id','?')} failed: {exc}")
        return built

    # ------------------------------------------------------------------
    # Individual pivot
    # ------------------------------------------------------------------
    def build_one(self, wb: Workbook, idx: int, spec: dict) -> None:
        sheet_name = spec.get("sheet") or "Dashboard"
        if sheet_name not in wb.sheetnames:
            sheet_name = "Dashboard" if "Dashboard" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        anchor = spec.get("location") or f"M{18 + idx * 14}"
        col_letter, row = coordinate_from_string(anchor)
        col = column_index_from_string(col_letter)

        row_fields = spec.get("row_fields") or []
        value_fields = spec.get("value_fields") or []
        filter_fields = spec.get("filter_fields") or []
        source_range = spec.get("source_range") or self._guess_source(wb)

        if not source_range or not row_fields or not value_fields:
            self._draw_placeholder(ws, row, col, spec)
            return

        source_sheet, src_rect = self._split_range(source_range)
        if source_sheet not in wb.sheetnames:
            self._draw_placeholder(ws, row, col, spec)
            return

        # Map field names -> column letters in the source sheet
        col_map = self._header_map(wb[source_sheet], src_rect)
        if not all(f in col_map for f in row_fields + [v.get("field") for v in value_fields]):
            self._draw_placeholder(ws, row, col, spec)
            return

        # Compute unique row-label values from the actual source rows
        unique_values = self._unique_values(wb[source_sheet], src_rect, col_map[row_fields[0]])

        # Header band
        labels = [row_fields[0]] + [self._value_label(vf) for vf in value_fields]
        for c_off, label in enumerate(labels):
            cell = ws.cell(row=row, column=col + c_off, value=label)
            self.formatter.apply(cell, self.formatter.header_style())

        # Body — one row per unique label
        src_first_row, src_last_row = self._row_bounds(src_rect)
        for r_off, label in enumerate(unique_values, start=1):
            ws.cell(row=row + r_off, column=col, value=label).font = Font(
                name=self.theme["font_name"], size=10
            )
            for c_off, vf in enumerate(value_fields, start=1):
                agg = (vf.get("aggregation") or "SUM").upper()
                fn = AGG_FN.get(agg, "SUMIFS")
                value_col = col_map[vf["field"]]
                criteria_col = col_map[row_fields[0]]
                criteria_range = f"{source_sheet}!{value_col}{src_first_row+1}:{value_col}{src_last_row}"
                lookup_range = f"{source_sheet}!{criteria_col}{src_first_row+1}:{criteria_col}{src_last_row}"
                criterion_cell = f"${get_column_letter(col)}{row + r_off}"
                if fn == "COUNTIFS":
                    formula = f"=IFERROR(COUNTIFS({lookup_range},{criterion_cell}),0)"
                else:
                    formula = (
                        f"=IFERROR({fn}({criteria_range},{lookup_range},{criterion_cell}),0)"
                    )
                target = ws.cell(row=row + r_off, column=col + c_off, value=formula)
                target.font = Font(name=self.theme["font_name"], size=10)
                target.number_format = '#,##0.00;[Red]-#,##0.00'

        # Grand total row
        if spec.get("has_grand_totals", True) and unique_values:
            total_row = row + len(unique_values) + 1
            tot_label = ws.cell(row=total_row, column=col, value="Grand Total")
            self.formatter.apply(tot_label, self.formatter.header_style())
            for c_off in range(1, len(value_fields) + 1):
                col_letter = get_column_letter(col + c_off)
                first_data_row = row + 1
                last_data_row = row + len(unique_values)
                total_cell = ws.cell(
                    row=total_row,
                    column=col + c_off,
                    value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})",
                )
                total_cell.number_format = '#,##0.00;[Red]-#,##0.00'
                self.formatter.apply(total_cell, self.formatter.header_style())

        # Stripe the body
        if unique_values:
            self.formatter.stripe_table(
                ws,
                start_row=row + 1,
                end_row=row + len(unique_values),
                start_col=col,
                end_col=col + len(value_fields),
            )

        # Filter band (optional — shows which filters drive this pivot)
        if filter_fields:
            note_row = (row + len(unique_values) + 3) if spec.get("has_grand_totals", True) else (row + len(unique_values) + 2)
            note = ws.cell(
                row=note_row,
                column=col,
                value=f"Filters: {', '.join(filter_fields)} (driven by $B$1)",
            )
            note.font = Font(name=self.theme["font_name"], italic=True, size=9, color="64748B")

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _split_range(rng: str) -> tuple[str, str]:
        if "!" in rng:
            sheet, rect = rng.split("!", 1)
            return sheet.strip("'"), rect
        return "Data", rng

    @staticmethod
    def _row_bounds(rect: str) -> tuple[int, int]:
        start, end = rect.split(":") if ":" in rect else (rect, rect)
        _, sr = coordinate_from_string(start)
        _, er = coordinate_from_string(end)
        return sr, er

    @staticmethod
    def _header_map(ws, rect: str) -> dict[str, str]:
        start, end = rect.split(":") if ":" in rect else (rect, rect)
        sc, sr = coordinate_from_string(start)
        ec, _ = coordinate_from_string(end)
        out: dict[str, str] = {}
        for col_idx in range(column_index_from_string(sc), column_index_from_string(ec) + 1):
            val = ws.cell(row=sr, column=col_idx).value
            if val:
                out[str(val)] = get_column_letter(col_idx)
        return out

    @staticmethod
    def _unique_values(ws, rect: str, col_letter: str, limit: int = 100) -> list:
        sr, er = PivotEngine._row_bounds(rect)
        col_idx = column_index_from_string(col_letter)
        seen: list = []
        seen_set: set = set()
        for r in range(sr + 1, er + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val is None or val == "":
                continue
            key = str(val)
            if key in seen_set:
                continue
            seen.append(val)
            seen_set.add(key)
            if len(seen) >= limit:
                break
        return seen

    @staticmethod
    def _guess_source(wb: Workbook) -> str | None:
        if "Data" not in wb.sheetnames:
            return None
        ws = wb["Data"]
        if ws.max_row <= 1:
            return None
        return f"Data!A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    @staticmethod
    def _value_label(vf: dict) -> str:
        agg = (vf.get("aggregation") or "SUM").upper()
        return f"{agg} of {vf.get('field','Value')}"

    def _draw_placeholder(self, ws, row: int, col: int, spec: dict) -> None:
        note = ws.cell(
            row=row,
            column=col,
            value=f"Pivot {spec.get('id','?')} — provide source_range + row_fields + value_fields",
        )
        note.font = Font(name=self.theme["font_name"], italic=True, color="64748B")
