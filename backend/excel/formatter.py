"""
FormatEngine — applies fonts, fills, borders, alignment, and all conditional
formatting rules described in the audit JSON. Also applies the theme.
"""
from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule, CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter


@dataclass
class FormatEngine:
    theme: dict

    # ------------------------------------------------------------------
    # Style helpers
    # ------------------------------------------------------------------
    def header_style(self):
        return {
            "font": Font(name=self.theme["font_name"], color=self.theme["header_fg"], bold=True, size=12),
            "fill": PatternFill("solid", fgColor=self.theme["header_bg"]),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": self.border(),
        }

    def title_style(self):
        return {
            "font": Font(name=self.theme["font_name"], color=self.theme["header_bg"], bold=True, size=20),
            "alignment": Alignment(horizontal="left", vertical="center"),
        }

    def kpi_label_style(self):
        return {
            "font": Font(name=self.theme["font_name"], color=self.theme["kpi_fg"], bold=True, size=10),
            "fill": PatternFill("solid", fgColor=self.theme["kpi_bg"]),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": self.border(),
        }

    def kpi_value_style(self):
        return {
            "font": Font(name=self.theme["font_name"], color=self.theme["kpi_fg"], bold=True, size=22),
            "fill": PatternFill("solid", fgColor=self.theme["kpi_bg"]),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": self.border(),
        }

    def body_style(self, alt: bool = False):
        return {
            "font": Font(name=self.theme["font_name"], size=10),
            "fill": PatternFill("solid", fgColor=self.theme["row_alt"]) if alt else None,
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": self.border(),
        }

    def border(self):
        side = Side(style="thin", color=self.theme["border_color"])
        return Border(left=side, right=side, top=side, bottom=side)

    @staticmethod
    def apply(cell, style: dict) -> None:
        for attr, val in (style or {}).items():
            if val is not None:
                setattr(cell, attr, val)

    # ------------------------------------------------------------------
    # Conditional formatting
    # ------------------------------------------------------------------
    def apply_conditional_formatting(self, ws, rules: list[dict]) -> int:
        applied = 0
        for rule in rules or []:
            rng = rule.get("range")
            if not rng:
                continue
            rtype = (rule.get("rule_type") or "color_scale").lower()
            try:
                if rtype == "color_scale":
                    ws.conditional_formatting.add(
                        rng,
                        ColorScaleRule(
                            start_type="min", start_color="F87171",
                            mid_type="percentile", mid_value=50, mid_color="FBBF24",
                            end_type="max", end_color="16A34A",
                        ),
                    )
                elif rtype == "data_bar":
                    ws.conditional_formatting.add(
                        rng,
                        DataBarRule(start_type="min", end_type="max", color=self.theme["accent1"]),
                    )
                elif rtype == "icon_set":
                    ws.conditional_formatting.add(
                        rng,
                        IconSetRule("3TrafficLights1", "percent", [0, 33, 67]),
                    )
                elif rtype == "cell_value":
                    op = rule.get("condition", "greaterThan") or "greaterThan"
                    val = rule.get("value", 0)
                    fmt = rule.get("format") or {}
                    diff = DifferentialStyle(
                        font=Font(color=fmt.get("font_color", "FFFFFF"), bold=bool(fmt.get("bold", False))),
                        fill=PatternFill("solid", fgColor=fmt.get("bg_color", self.theme["accent1"])),
                    )
                    cell_rule = CellIsRule(operator=op, formula=[str(val)], dxf=diff)
                    ws.conditional_formatting.add(rng, cell_rule)
                applied += 1
            except Exception:
                continue
        return applied

    # ------------------------------------------------------------------
    # Bulk row striping
    # ------------------------------------------------------------------
    def stripe_table(self, ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                if (r - start_row) % 2 == 1:
                    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=self.theme["row_alt"])

    # ------------------------------------------------------------------
    # Sheet-level theming
    # ------------------------------------------------------------------
    def theme_sheet(self, ws, hide_gridlines: bool = True) -> None:
        ws.sheet_view.showGridLines = not hide_gridlines
        ws.sheet_properties.tabColor = self.theme["tab_color"]
        for col_idx in range(1, 30):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16
