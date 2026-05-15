"""
ChartEngine — builds every chart type from the audit JSON.

Supports: bar, line, pie, donut, area, scatter, combo, radar, bubble, gauge,
bullet, cylinder, nested_donut, waterfall, funnel, timeline, lollipop,
progress_bar, smooth_line, stacked_bar, grouped_bar.

Where openpyxl can't natively express a chart type, we approximate using
the closest base type with appropriate styling (gauge → donut, waterfall →
stacked bar, funnel → horizontal bar decreasing, etc.).
"""
from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from openpyxl.chart import (
    AreaChart,
    BarChart,
    BarChart3D,
    BubbleChart,
    DoughnutChart,
    LineChart,
    PieChart,
    RadarChart,
    Reference,
    ScatterChart,
    Series,
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.fill import ColorChoice as FillColorChoice, PatternFillProperties, SolidColorFillProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from loguru import logger


@dataclass
class ChartEngine:
    theme: dict

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------
    def build_all(self, wb, charts: list[dict]) -> int:
        built = 0
        for spec in charts or []:
            try:
                self.build_one(wb, spec)
                built += 1
            except Exception as exc:
                logger.warning(f"chart {spec.get('id','?')} failed: {exc}")
        return built

    def build_one(self, wb, spec: dict) -> None:
        ctype = (spec.get("type") or "bar").lower()
        builder = self._dispatch.get(ctype, self._build_bar)
        target_sheet = spec.get("sheet") or "Dashboard"
        if target_sheet not in wb.sheetnames:
            target_sheet = "Dashboard" if "Dashboard" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[target_sheet]
        chart = builder(wb, spec)
        if chart is None:
            return
        # Title (formula-driven if title starts with =)
        title = spec.get("title") or ""
        if title:
            chart.title = title
        # Legend & labels
        chart.legend = None if not spec.get("has_legend", True) else chart.legend
        if spec.get("has_data_labels"):
            chart.dataLabels = DataLabelList(showVal=True)
        # Axis titles
        x_title = spec.get("axis_x_title")
        y_title = spec.get("axis_y_title")
        if hasattr(chart, "x_axis") and x_title and chart.x_axis is not None:
            chart.x_axis.title = x_title
        if hasattr(chart, "y_axis") and y_title and chart.y_axis is not None:
            chart.y_axis.title = y_title
        # Anchor
        anchor = spec.get("location") or "B10"
        chart.width = float(spec.get("width", 16))
        chart.height = float(spec.get("height", 9))
        ws.add_chart(chart, anchor)

    # ------------------------------------------------------------------
    # Series helpers
    # ------------------------------------------------------------------
    def _add_series(self, chart, wb, spec: dict) -> None:
        for idx, series in enumerate(spec.get("series", [])):
            try:
                values_ref = self._reference(wb, series.get("values_range"))
                if values_ref is None:
                    continue
                s = Series(values=values_ref, title=series.get("name") or f"Series {idx+1}")
                cats_ref = self._reference(wb, series.get("categories_range"))
                if cats_ref is not None:
                    s.categories = cats_ref
                color = (series.get("color") or self.theme["chart_colors"][idx % len(self.theme["chart_colors"])]).lstrip("#")
                gp = GraphicalProperties(solidFill=color)
                gp.line = LineProperties(solidFill=color)
                s.graphicalProperties = gp
                chart.series.append(s)
            except Exception as exc:
                logger.warning(f"series {idx} failed: {exc}")

    @staticmethod
    def _reference(wb, range_str: str | None):
        if not range_str:
            return None
        try:
            if "!" in range_str:
                sheet_name, cell_range = range_str.split("!", 1)
                sheet_name = sheet_name.strip("'")
            else:
                sheet_name, cell_range = "Data", range_str
            if sheet_name not in wb.sheetnames:
                return None
            ws = wb[sheet_name]
            if ":" not in cell_range:
                cell_range = f"{cell_range}:{cell_range}"
            start, end = cell_range.split(":")
            sc, sr = coordinate_from_string(start)
            ec, er = coordinate_from_string(end)
            return Reference(ws, min_col=column_index_from_string(sc), min_row=sr,
                             max_col=column_index_from_string(ec), max_row=er)
        except Exception:
            return None

    # ------------------------------------------------------------------
    # Builders for each chart type
    # ------------------------------------------------------------------
    def _build_bar(self, wb, spec):
        chart = BarChart()
        chart.type = "col"
        self._add_series(chart, wb, spec)
        return chart

    def _build_stacked_bar(self, wb, spec):
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100
        self._add_series(chart, wb, spec)
        return chart

    def _build_grouped_bar(self, wb, spec):
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        self._add_series(chart, wb, spec)
        return chart

    def _build_horizontal_bar(self, wb, spec):
        chart = BarChart()
        chart.type = "bar"
        self._add_series(chart, wb, spec)
        return chart

    def _build_line(self, wb, spec):
        chart = LineChart()
        self._add_series(chart, wb, spec)
        return chart

    def _build_smooth_line(self, wb, spec):
        chart = LineChart()
        self._add_series(chart, wb, spec)
        for s in chart.series:
            s.smooth = True
        return chart

    def _build_area(self, wb, spec):
        chart = AreaChart()
        self._add_series(chart, wb, spec)
        return chart

    def _build_pie(self, wb, spec):
        chart = PieChart()
        self._add_series(chart, wb, spec)
        return chart

    def _build_donut(self, wb, spec):
        chart = DoughnutChart()
        chart.holeSize = 50
        self._add_series(chart, wb, spec)
        return chart

    def _build_nested_donut(self, wb, spec):
        chart = DoughnutChart()
        chart.holeSize = 30
        self._add_series(chart, wb, spec)
        return chart

    def _build_scatter(self, wb, spec):
        chart = ScatterChart()
        chart.style = 13
        for idx, series in enumerate(spec.get("series", [])):
            xref = self._reference(wb, series.get("categories_range"))
            yref = self._reference(wb, series.get("values_range"))
            if xref is None or yref is None:
                continue
            s = Series(values=yref, x_values=xref, title=series.get("name") or f"Series {idx+1}")
            chart.series.append(s)
        return chart

    def _build_radar(self, wb, spec):
        chart = RadarChart()
        chart.type = "filled"
        chart.style = 26
        self._add_series(chart, wb, spec)
        return chart

    def _build_bubble(self, wb, spec):
        chart = BubbleChart()
        chart.style = 18
        self._add_series(chart, wb, spec)
        return chart

    def _build_combo(self, wb, spec):
        bar = BarChart()
        line = LineChart()
        series_list = spec.get("series", [])
        for idx, series in enumerate(series_list):
            sub = (series.get("chart_type") or ("line" if idx else "bar")).lower()
            target = line if sub in {"line", "smooth_line", "trend"} else bar
            ref = self._reference(wb, series.get("values_range"))
            if ref is None:
                continue
            s = Series(values=ref, title=series.get("name") or f"Series {idx+1}")
            cats = self._reference(wb, series.get("categories_range"))
            if cats is not None:
                s.categories = cats
            target.series.append(s)
        bar += line  # combine
        return bar

    def _build_cylinder(self, wb, spec):
        chart = BarChart3D()
        chart.shape = "cylinder"
        self._add_series(chart, wb, spec)
        return chart

    def _build_gauge(self, wb, spec):
        # Approximated as half-donut: value + remainder.
        chart = DoughnutChart()
        chart.holeSize = 60
        chart.firstSliceAng = 270
        self._add_series(chart, wb, spec)
        return chart

    def _build_bullet(self, wb, spec):
        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "stacked"
        chart.overlap = 100
        self._add_series(chart, wb, spec)
        return chart

    def _build_waterfall(self, wb, spec):
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100
        self._add_series(chart, wb, spec)
        return chart

    def _build_funnel(self, wb, spec):
        chart = BarChart()
        chart.type = "bar"
        self._add_series(chart, wb, spec)
        return chart

    def _build_timeline(self, wb, spec):
        return self._build_scatter(wb, spec)

    def _build_lollipop(self, wb, spec):
        chart = ScatterChart()
        chart.style = 13
        for idx, series in enumerate(spec.get("series", [])):
            xref = self._reference(wb, series.get("categories_range"))
            yref = self._reference(wb, series.get("values_range"))
            if xref is None or yref is None:
                continue
            s = Series(values=yref, x_values=xref, title=series.get("name") or f"Series {idx+1}")
            chart.series.append(s)
        return chart

    def _build_progress(self, wb, spec):
        return self._build_stacked_bar(wb, spec)

    @property
    def _dispatch(self):
        return {
            "bar": self._build_bar,
            "column": self._build_bar,
            "grouped_bar": self._build_grouped_bar,
            "stacked_bar": self._build_stacked_bar,
            "horizontal_bar": self._build_horizontal_bar,
            "line": self._build_line,
            "smooth_line": self._build_smooth_line,
            "area": self._build_area,
            "pie": self._build_pie,
            "donut": self._build_donut,
            "doughnut": self._build_donut,
            "nested_donut": self._build_nested_donut,
            "scatter": self._build_scatter,
            "radar": self._build_radar,
            "spider": self._build_radar,
            "bubble": self._build_bubble,
            "combo": self._build_combo,
            "cylinder": self._build_cylinder,
            "gauge": self._build_gauge,
            "speedometer": self._build_gauge,
            "bullet": self._build_bullet,
            "waterfall": self._build_waterfall,
            "funnel": self._build_funnel,
            "timeline": self._build_timeline,
            "gantt": self._build_timeline,
            "lollipop": self._build_lollipop,
            "progress_bar": self._build_progress,
        }
