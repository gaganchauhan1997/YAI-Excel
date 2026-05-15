"""
ExcelParser — read .xlsx/.xls into a DataFrame and inspect existing
structure (sheet count, charts present, named ranges).
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from loguru import logger


@dataclass
class ExcelParser:
    def read(self, path: str | Path, sheet: str | int | None = 0) -> pd.DataFrame:
        return pd.read_excel(path, sheet_name=sheet)

    def read_all_sheets(self, path: str | Path) -> dict[str, pd.DataFrame]:
        return pd.read_excel(path, sheet_name=None)

    def inspect(self, path: str | Path) -> dict[str, Any]:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        info = {
            "sheets": wb.sheetnames,
            "sheet_count": len(wb.sheetnames),
            "named_ranges": list(getattr(wb.defined_names, "definedName", []) or []),
            "charts_per_sheet": {},
        }
        for s in wb.sheetnames:
            ws = wb[s]
            info["charts_per_sheet"][s] = len(getattr(ws, "_charts", []) or [])
        logger.info(f"ExcelParser.inspect | sheets={info['sheets']} charts={info['charts_per_sheet']}")
        return info
