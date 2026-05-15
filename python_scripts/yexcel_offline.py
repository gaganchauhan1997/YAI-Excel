"""
YAI-Excel — Offline Reference Implementation (openpyxl)
=======================================================

This is the *learning copy* of the production Cloudflare Worker.
Worker.js does the same thing in pure JavaScript so it can run on Cloudflare's
edge for free — this Python version is here for you to read, run locally,
and modify before / instead of asking the AI to do the work.

70% of the intelligence lives here in Python (deterministic):
  • CSV parsing
  • Numeric / categorical column detection
  • SUM / AVG / MIN / MAX / GROUP-BY aggregations
  • KPI selection by impact
  • Chart dimension/metric picking
  • Excel/XLSX generation with embedded charts (BarChart, DoughnutChart, PieChart, stacked)

30% is reserved for AI (optional — graceful fallback if no key):
  • Title generation
  • Domain naming
  • KPI label polishing (e.g. "SalesRevenue" → "Sales Revenue")
  • Smart chart titles ("Revenue vs Cost by Product")

Run:
    pip install openpyxl
    python yexcel_offline.py path/to/data.csv [theme]

Themes: midnight emerald crimson slate amber ocean violet rose carbon arctic
"""

import csv
import re
import sys
from pathlib import Path
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, DoughnutChart, Reference
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

THEMES = {
    'midnight': dict(bg='0A0A0F', card='12121A', accent='00F0FF', text='E0E0E0', header='1A1A2E',
                     chart=['00F0FF', '7C3AED', '10B981', 'F59E0B', 'EF4444']),
    'emerald':  dict(bg='022C22', card='064E3B', accent='10B981', text='D1FAE5', header='064E3B',
                     chart=['10B981', '34D399', '6EE7B7', 'A7F3D0', '059669']),
    'crimson':  dict(bg='1A0505', card='2A0A0A', accent='EF4444', text='FECACA', header='2A0A0A',
                     chart=['EF4444', 'F87171', 'FCA5A5', 'F43F5E', 'E11D48']),
    'slate':    dict(bg='0F172A', card='1E293B', accent='94A3B8', text='E2E8F0', header='1E293B',
                     chart=['94A3B8', '60A5FA', 'A78BFA', '34D399', 'F59E0B']),
    'amber':    dict(bg='1C0A00', card='2D1500', accent='F59E0B', text='FEF3C7', header='2D1500',
                     chart=['F59E0B', 'FBBF24', 'FCD34D', 'EF4444', '10B981']),
    'ocean':    dict(bg='0C1A2E', card='0C4A6E', accent='0EA5E9', text='E0F2FE', header='0C4A6E',
                     chart=['0EA5E9', '38BDF8', '7DD3FC', '06B6D4', '0284C7']),
    'violet':   dict(bg='1A0533', card='2E1065', accent='8B5CF6', text='EDE9FE', header='2E1065',
                     chart=['8B5CF6', 'A78BFA', 'C4B5FD', '7C3AED', '6D28D9']),
    'rose':     dict(bg='1A0010', card='4C0519', accent='F43F5E', text='FFE4E6', header='4C0519',
                     chart=['F43F5E', 'FB7185', 'FDA4AF', 'E11D48', 'F59E0B']),
    'carbon':   dict(bg='0D0D0D', card='1A1A1A', accent='6B7280', text='F3F4F6', header='1A1A1A',
                     chart=['6B7280', '9CA3AF', 'D1D5DB', '4B5563', '374151']),
    'arctic':   dict(bg='050E2A', card='0C1445', accent='60A5FA', text='DBEAFE', header='0C1445',
                     chart=['60A5FA', '93C5FD', '38BDF8', '818CF8', '34D399']),
}

IDENTIFIER_RE = re.compile(
    r'^(year|month|day|id|code|sku|isbn|upc|barcode|zip|pincode|postal|phone|fax|account_?no|order_?id|customer_?id|employee_?id|invoice|reference)$|'
    r'(_year|_month|_id|_code|_no|_number)$',
    re.IGNORECASE,
)


def parse_csv(path: Path):
    rows = []
    with path.open(newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        for r in reader:
            rows.append({h: (r.get(h) or '').strip() for h in headers})
    return headers, rows


def to_num(v):
    """Strip currency / percent / Indian symbols and try float."""
    s = re.sub(r'[$,€£₹%\s]', '', str(v or ''))
    try:
        return float(s)
    except ValueError:
        return None


def compute_stats(rows, headers):
    """Classify each column as numeric or categorical and compute summary stats."""
    stats = {}
    for h in headers:
        values = [r[h] for r in rows if r[h] not in ('', None)]
        numerics = [n for n in (to_num(v) for v in values) if n is not None]
        is_numeric = len(numerics) >= len(values) * 0.6 and values
        if is_numeric:
            stats[h] = dict(
                type='numeric',
                sum=round(sum(numerics), 2),
                avg=round(sum(numerics) / len(numerics), 2),
                min=min(numerics), max=max(numerics), count=len(numerics),
            )
        else:
            counts = Counter(values)
            top = dict(counts.most_common(20))
            stats[h] = dict(type='categorical', unique_count=len(counts),
                            value_counts=top, sample=list(top.keys())[:8])
    return stats


def group_by(rows, headers, stats):
    """Compute SUM(metric) for every (dim, dim_value) pair — the same logic Excel SUMIFS uses."""
    # Promote identifier-shaped numeric columns to dimensions so Year / Month work
    for h in headers:
        if stats[h]['type'] == 'numeric' and IDENTIFIER_RE.search(h):
            counts = Counter(str(r[h]) for r in rows if r[h] != '')
            stats[h]['value_counts'] = dict(counts.most_common(20))
            stats[h]['unique_count'] = len(counts)
            stats[h]['sample'] = list(counts.keys())[:8]

    dimensions = [h for h in headers if
                  (stats[h]['type'] == 'categorical' and stats[h]['unique_count'] <= 24) or
                  (stats[h]['type'] == 'numeric' and IDENTIFIER_RE.search(h) and stats[h]['unique_count'] <= 24)]
    metrics = [h for h in headers if stats[h]['type'] == 'numeric' and not IDENTIFIER_RE.search(h)]

    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for r in rows:
        for d in dimensions:
            dv = str(r[d])
            for m in metrics:
                v = to_num(r[m])
                if v is not None:
                    grouped[d][dv][m] += v
    return dimensions, metrics, grouped


def cross_agg(rows, dimA, dimB, metric, statsA, statsB):
    """Two-dimensional pivot for stacked-bar chart 4."""
    aVals = list(statsA['value_counts'].keys())[:12]
    bVals = list(statsB['value_counts'].keys())[:12]
    matrix = {a: {b: 0.0 for b in bVals} for a in aVals}
    for r in rows:
        a, b = str(r.get(dimA, '')), str(r.get(dimB, ''))
        if a not in matrix or b not in matrix[a]:
            continue
        v = to_num(r[metric])
        if v is not None:
            matrix[a][b] += v
    return aVals, bVals, matrix


def guess_domain(headers):
    s = ' '.join(headers).lower()
    if re.search(r'production|plant|defect|supervisor|shift|manufactur', s): return 'Manufacturing'
    if re.search(r'sale|revenue|customer|order|product', s): return 'Sales'
    if re.search(r'employee|salary|hire|department', s): return 'HR'
    if re.search(r'budget|expense|account|profit', s): return 'Finance'
    if re.search(r'student|grade|course|teacher', s): return 'Education'
    return 'Business'


def build_dashboard(csv_path: Path, theme: str = 'midnight', out_path: Path = None):
    headers, rows = parse_csv(csv_path)
    if not rows:
        raise ValueError("CSV is empty")
    stats = compute_stats(rows, headers)
    dimensions, metrics, grouped = group_by(rows, headers, stats)
    domain = guess_domain(headers)
    title = f"{domain} Dashboard"
    tc = THEMES.get(theme, THEMES['midnight'])

    # ── KPI picks (deterministic — top 5 by sum, identifier columns excluded) ──
    kpi_cols = sorted(metrics, key=lambda m: stats[m]['sum'], reverse=True)[:5]
    kpis = []
    for col in kpi_cols:
        lower = col.lower()
        if re.search(r'revenue|sales|cost|price|amount|profit', lower):
            fmt = 'currency'
        elif re.search(r'efficien|percent|rate|ratio|margin', lower):
            fmt = 'percent'
        elif stats[col]['sum'] > 100000:
            fmt = 'compact'
        else:
            fmt = 'number'
        kpis.append(dict(
            label=re.sub(r'(?<!^)(?=[A-Z])', ' ', col).replace('_', ' ').title(),
            value=stats[col]['sum'], format=fmt, column=col,
        ))

    # ── Chart picks (deterministic) ──
    def pick_dim(idx=0): return dimensions[idx] if idx < len(dimensions) else dimensions[0]
    def pick_metric(idx=0): return metrics[idx] if idx < len(metrics) else metrics[0]

    monthish = next((d for d in dimensions if re.search(r'month|date|quarter|period|week', d, re.I)), None)
    month_idx = dimensions.index(monthish) if monthish else 0

    c1 = dict(dim=pick_dim(0), metric=pick_metric(0), kind='hbar')
    c2 = dict(dim=pick_dim(1), metric=pick_metric(0), kind='doughnut')
    c3 = dict(dim=pick_dim(2 if len(dimensions) > 2 else 0), metric=pick_metric(1 if len(metrics) > 1 else 0), kind='pie')

    # chart4 stacked bar (Month × Product × Revenue style)
    c4DimA = pick_dim(month_idx)
    c4DimB = next((d for d in dimensions if d != c4DimA), c4DimA)
    c4Metric = pick_metric(0)
    aVals, bVals, matrix = cross_agg(rows, c4DimA, c4DimB, c4Metric, stats[c4DimA], stats[c4DimB])

    # ── Build XLSX ──
    wb = Workbook()
    dash = wb.active
    dash.title = 'Dashboard'
    dash.sheet_view.showGridLines = False
    dash.sheet_properties.tabColor = tc['accent']

    # Title banner
    dash['A1'] = f"{title.upper()} — IN EXCEL"
    dash.merge_cells('A1:P2')
    for row in dash['A1:P2']:
        for cell in row:
            cell.fill = PatternFill('solid', fgColor=tc['header'])
            cell.font = Font(bold=True, size=22, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Tab strip
    tabs = ['Overview', 'Production Analysis', 'Quality Analysis', 'Cost & Revenue', 'Efficiency & Downtime']
    tab_starts = ['A3', 'D3', 'G3', 'J3', 'M3']
    for i, t in enumerate(tabs):
        dash[tab_starts[i]] = t
    for row in dash['A3:P3']:
        for cell in row:
            cell.fill = PatternFill('solid', fgColor=tc['header'])
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Page background
    for r in range(4, 45):
        for c in range(1, 17):
            cell = dash.cell(row=r, column=c)
            if not cell.fill or cell.fill.fgColor.rgb is None:
                cell.fill = PatternFill('solid', fgColor=tc['bg'])

    # KPI cards (D5:M6, 2-col cards, 5 of them)
    kpi_border = Border(*(Side(style='thin', color=tc['accent']),) * 4)
    accent_font_big = Font(bold=True, size=26, color=tc['accent'])
    label_font = Font(size=10, color='B3B3B3')

    for i, kpi in enumerate(kpis):
        c0 = 4 + i * 2  # col 4,6,8,10,12 (1-based)
        # label row 5
        label_cell = dash.cell(row=5, column=c0, value=kpi['label'].upper())
        label_cell.fill = PatternFill('solid', fgColor=tc['card'])
        label_cell.font = label_font
        label_cell.alignment = Alignment(horizontal='center', vertical='center')
        label_cell.border = kpi_border
        dash.merge_cells(start_row=5, start_column=c0, end_row=5, end_column=c0 + 1)

        # value row 6
        val_cell = dash.cell(row=6, column=c0, value=kpi['value'])
        val_cell.fill = PatternFill('solid', fgColor=tc['card'])
        val_cell.font = accent_font_big
        val_cell.alignment = Alignment(horizontal='center', vertical='center')
        val_cell.border = kpi_border
        if kpi['format'] == 'currency':
            val_cell.number_format = '"$"#,##0.0,,"M"'
        elif kpi['format'] == 'percent':
            val_cell.number_format = '0.00"%"'
        else:
            val_cell.number_format = '#,##0.0,"K"'
        dash.merge_cells(start_row=6, start_column=c0, end_row=6, end_column=c0 + 1)

    dash.row_dimensions[5].height = 18
    dash.row_dimensions[6].height = 32

    # Left filter (col A:B)
    if dimensions:
        f1 = dimensions[0]
        dash['A5'] = f1.upper()
        dash.merge_cells('A5:B5')
        dash['A5'].font = Font(bold=True, size=14, color='FFFFFF')
        dash['A5'].fill = PatternFill('solid', fgColor=tc['card'])
        for i, v in enumerate(list(stats[f1]['value_counts'].keys())[:8]):
            r = 6 + i // 2
            c = 1 + i % 2
            cell = dash.cell(row=r, column=c, value=v)
            cell.fill = PatternFill('solid', fgColor=tc['card'])
            cell.font = Font(color='FFFFFF', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    # Right filter (col O:P)
    if len(dimensions) > 1:
        f2 = dimensions[1]
        dash['O5'] = f2.upper()
        dash.merge_cells('O5:P5')
        dash['O5'].font = Font(bold=True, size=14, color='FFFFFF')
        dash['O5'].fill = PatternFill('solid', fgColor=tc['card'])
        for i, v in enumerate(list(stats[f2]['value_counts'].keys())[:8]):
            r = 6 + i // 2
            c = 15 + i % 2
            cell = dash.cell(row=r, column=c, value=v)
            cell.fill = PatternFill('solid', fgColor=tc['card'])
            cell.font = Font(color='FFFFFF', size=10)

    # Column widths
    dash.column_dimensions['A'].width = 14
    dash.column_dimensions['B'].width = 14
    dash.column_dimensions['C'].width = 3
    for col_letter in 'DEFGHIJKLM':
        dash.column_dimensions[col_letter].width = 13
    dash.column_dimensions['N'].width = 3
    dash.column_dimensions['O'].width = 14
    dash.column_dimensions['P'].width = 14

    # Hidden chart data ranges
    def write_chart_data(start_row, label_col, value_cols, headers_row, label_values, data_rows):
        for c, h in enumerate(headers_row):
            cell = dash.cell(row=start_row, column=label_col + c, value=h)
            cell.font = Font(bold=True, color='111827')
        for i, label in enumerate(label_values):
            dash.cell(row=start_row + 1 + i, column=label_col, value=label)
            for c, v in enumerate(data_rows[i]):
                dash.cell(row=start_row + 1 + i, column=label_col + 1 + c, value=v)

    # Chart 1: hbar
    c1_labels = list(stats[c1['dim']]['value_counts'].keys())[:10]
    c1_data = [[grouped[c1['dim']][l][c1['metric']]] for l in c1_labels]
    write_chart_data(50, 2, 1, [c1['dim'], c1['metric']], c1_labels, c1_data)

    # Chart 2: doughnut
    c2_labels = list(stats[c2['dim']]['value_counts'].keys())[:6]
    c2_data = [[grouped[c2['dim']][l][c2['metric']]] for l in c2_labels]
    write_chart_data(70, 2, 1, [c2['dim'], c2['metric']], c2_labels, c2_data)

    # Chart 3: pie
    c3_labels = list(stats[c3['dim']]['value_counts'].keys())[:6]
    c3_data = [[grouped[c3['dim']][l][c3['metric']]] for l in c3_labels]
    write_chart_data(90, 2, 1, [c3['dim'], c3['metric']], c3_labels, c3_data)

    # Chart 4: stacked
    c4_headers = [c4DimA] + bVals
    c4_data = [[matrix[a][b] for b in bVals] for a in aVals]
    write_chart_data(110, 2, len(bVals), c4_headers, aVals, c4_data)

    # Build chart objects
    # Chart 1: Horizontal Bar
    bar1 = BarChart()
    bar1.type = 'bar'
    bar1.style = 10
    bar1.title = f"{c1['metric']} by {c1['dim']}"
    bar1.y_axis.title = c1['dim']
    bar1.x_axis.title = c1['metric']
    data_ref = Reference(dash, min_col=3, min_row=50, max_row=50 + len(c1_labels))
    cats_ref = Reference(dash, min_col=2, min_row=51, max_row=50 + len(c1_labels))
    bar1.add_data(data_ref, titles_from_data=True)
    bar1.set_categories(cats_ref)
    dash.add_chart(bar1, 'D10')

    # Chart 2: Doughnut
    dn = DoughnutChart()
    dn.title = f"{c2['metric']} by {c2['dim']}"
    data_ref = Reference(dash, min_col=3, min_row=70, max_row=70 + len(c2_labels))
    cats_ref = Reference(dash, min_col=2, min_row=71, max_row=70 + len(c2_labels))
    dn.add_data(data_ref, titles_from_data=True)
    dn.set_categories(cats_ref)
    dash.add_chart(dn, 'I10')

    # Chart 3: Pie
    pie = PieChart()
    pie.title = f"{c3['metric']} by {c3['dim']}"
    data_ref = Reference(dash, min_col=3, min_row=90, max_row=90 + len(c3_labels))
    cats_ref = Reference(dash, min_col=2, min_row=91, max_row=90 + len(c3_labels))
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats_ref)
    dash.add_chart(pie, 'N10')

    # Chart 4: Stacked column
    bar4 = BarChart()
    bar4.type = 'col'
    bar4.grouping = 'stacked'
    bar4.overlap = 100
    bar4.title = f"{c4Metric} by {c4DimA} and {c4DimB}"
    data_ref = Reference(dash, min_col=3, max_col=2 + len(bVals), min_row=110, max_row=110 + len(aVals))
    cats_ref = Reference(dash, min_col=2, min_row=111, max_row=110 + len(aVals))
    bar4.add_data(data_ref, titles_from_data=True)
    bar4.set_categories(cats_ref)
    dash.add_chart(bar4, 'D25')

    # ── Data sheet ──
    data_ws = wb.create_sheet('Data')
    data_ws.append(headers)
    for r in rows:
        data_ws.append([to_num(r[h]) if to_num(r[h]) is not None else r[h] for h in headers])

    # ── Excel Tutor sheet ──
    tutor = wb.create_sheet('Excel Tutor')
    tutor.sheet_view.showGridLines = False
    tutor.sheet_properties.tabColor = '22C55E'

    m1, m2 = metrics[0], (metrics[1] if len(metrics) > 1 else metrics[0])
    d1 = dimensions[0] if dimensions else headers[0]
    m1_col = get_column_letter(headers.index(m1) + 1)
    d1_col = get_column_letter(headers.index(d1) + 1)
    m2_col = get_column_letter(headers.index(m2) + 1)
    sample = next(iter(stats[d1]['value_counts'].keys()))

    LESSONS = [
        ('HEADER', '🎓 EXCEL TUTOR — LEARN FROM YOUR OWN DATA'),
        ('BLANK',),
        ('HEADER', 'LESSON 1 — TOTALS, AVERAGES, COUNTS'),
        ('PAIR', f'=SUM(Data!{m1_col}:{m1_col})',
                 f'Sums every value in the "{m1}" column on the Data sheet. SUM is your first reflex when you need a total.'),
        ('PAIR', f'=AVERAGE(Data!{m1_col}:{m1_col})',
                 f'Average of "{m1}". Use AVERAGE for mean, MEDIAN for middle.'),
        ('PAIR', '=COUNTA(Data!A:A)-1',
                 'Counts how many records you have (minus header). COUNTA = non-empty cells.'),
        ('TIP', '💡 Whole-column references (e.g. C:C) stay correct as data grows.'),
        ('BLANK',),
        ('HEADER', 'LESSON 2 — CONDITIONAL TOTALS WITH SUMIFS'),
        ('PAIR', f'=SUMIFS(Data!{m1_col}:{m1_col},Data!{d1_col}:{d1_col},"{sample}")',
                 f'Sums "{m1}" only when "{d1}" equals "{sample}". The workhorse formula behind every dashboard KPI.'),
        ('PAIR', f'=COUNTIFS(Data!{d1_col}:{d1_col},"{sample}")',
                 'Counts how many rows match the criterion.'),
        ('TIP', '💡 SUMIFS takes the SUM column FIRST, then pairs of (criteria_range, criteria).'),
        ('BLANK',),
        ('HEADER', 'LESSON 3 — LOOKUPS (XLOOKUP / INDEX-MATCH)'),
        ('PAIR', f'=XLOOKUP("FindMe",Data!{d1_col}:{d1_col},Data!{m1_col}:{m1_col},"Not Found")',
                 'XLOOKUP — modern, 4th arg is the not-found fallback.'),
        ('PAIR', f'=INDEX(Data!{m1_col}:{m1_col},MATCH("FindMe",Data!{d1_col}:{d1_col},0))',
                 'INDEX/MATCH — works everywhere, beats VLOOKUP every time.'),
        ('BLANK',),
        ('HEADER', 'LESSON 4 — KPIs WITH SMART FORMATTING'),
        ('PAIR', f'=TEXT(SUM(Data!{m1_col}:{m1_col}),"$#,##0.0,,""M""")',
                 'Wraps a number in millions with the M suffix. ,, divides by 1,000,000.'),
        ('PAIR', f'=ROUND(SUM(Data!{m1_col}:{m1_col})/SUM(Data!{m2_col}:{m2_col})*100,2)&"%"',
                 'Ratio between two metrics with % suffix.'),
        ('BLANK',),
        ('HEADER', 'LESSON 5 — PIVOTS WITHOUT PIVOT TABLES'),
        ('PAIR', f'=UNIQUE(Data!{d1_col}:{d1_col})',
                 f'Pulls every distinct value from "{d1}" — modern pivot dim list.'),
        ('PAIR', f'=SORT(UNIQUE(Data!{d1_col}:{d1_col}))', 'Same, sorted A→Z.'),
        ('PAIR', f'=FILTER(Data!A:Z,Data!{d1_col}:{d1_col}="{sample}")',
                 'Live slicer — only matching rows.'),
        ('TIP', '💡 Dynamic-array functions spill across cells. Put them in one cell and watch the rest auto-fill.'),
        ('BLANK',),
        ('HEADER', 'LESSON 6 — KEYBOARD SHORTCUTS'),
        ('PAIR', 'Ctrl + Shift + L', 'Toggle filters on/off.'),
        ('PAIR', 'Ctrl + T', 'Convert a range to a Table — auto-extends formulas.'),
        ('PAIR', 'F4', 'Toggle absolute/relative reference.'),
        ('PAIR', 'Ctrl + ;', 'Insert today\'s date as a hard value.'),
        ('BLANK',),
        ('HEADER', f'🎯 RECREATE THIS DASHBOARD ({domain.upper()} KPI FORMULAS)'),
    ]
    for k in kpis:
        col = get_column_letter(headers.index(k['column']) + 1)
        LESSONS.append(('PAIR', f'=SUM(Data!{col}:{col})', f'Recomputes the "{k["label"]}" KPI.'))
    LESSONS.append(('TIP', '🚀 Right-click any chart on the Dashboard → "Select Data" to see exactly which columns drive it.'))

    code_fill = PatternFill('solid', fgColor='F3F4F6')
    body_fill = PatternFill('solid', fgColor='FFFFFF')
    tip_fill = PatternFill('solid', fgColor='FFFBEB')
    header_fill = PatternFill('solid', fgColor=tc['header'])
    text_dark = Font(color='111827', size=11)
    text_white = Font(color='FFFFFF', size=14, bold=True)
    code_font = Font(name='Consolas', size=10, color='111827', bold=True)
    light_border = Border(*(Side(style='thin', color='D1D5DB'),) * 4)

    for idx, lesson in enumerate(LESSONS, start=1):
        kind = lesson[0]
        if kind == 'HEADER':
            tutor.cell(row=idx, column=1, value=lesson[1])
            tutor.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=7)
            for c in range(1, 8):
                cell = tutor.cell(row=idx, column=c)
                cell.fill = header_fill
                cell.font = text_white
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            tutor.row_dimensions[idx].height = 26
        elif kind == 'PAIR':
            tutor.cell(row=idx, column=1, value=lesson[1])
            tutor.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=3)
            tutor.cell(row=idx, column=4, value=lesson[2])
            tutor.merge_cells(start_row=idx, start_column=4, end_row=idx, end_column=7)
            for c in range(1, 4):
                cell = tutor.cell(row=idx, column=c)
                cell.fill = code_fill
                cell.font = code_font
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
                cell.border = light_border
            for c in range(4, 8):
                cell = tutor.cell(row=idx, column=c)
                cell.fill = body_fill
                cell.font = text_dark
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = light_border
            tutor.row_dimensions[idx].height = 28
        elif kind == 'TIP':
            tutor.cell(row=idx, column=1, value=lesson[1])
            tutor.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=7)
            for c in range(1, 8):
                cell = tutor.cell(row=idx, column=c)
                cell.fill = tip_fill
                cell.font = text_dark
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
                cell.border = light_border
            tutor.row_dimensions[idx].height = 28
        else:
            tutor.row_dimensions[idx].height = 8

    for col_letter in 'ABC':
        tutor.column_dimensions[col_letter].width = 30
    for col_letter in 'DEFG':
        tutor.column_dimensions[col_letter].width = 40

    # Save
    out_path = out_path or csv_path.with_suffix('.xlsx')
    wb.save(out_path)
    return out_path


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("\nUsage: python yexcel_offline.py <csv-path> [theme]")
        sys.exit(1)
    csv_path = Path(sys.argv[1])
    theme = sys.argv[2] if len(sys.argv) > 2 else 'midnight'
    out = build_dashboard(csv_path, theme)
    print(f"✓ Wrote dashboard → {out}")


if __name__ == '__main__':
    main()
